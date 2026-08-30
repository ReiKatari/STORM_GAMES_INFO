"""
STORM GAMES INFO - Application for collecting game information from various sources
Main GUI application with dark theme and Caching support
"""
import sys
import os
import requests
import time
import shutil
import subprocess
from io import BytesIO
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QComboBox, QPushButton, QListWidget, QListWidgetItem,
    QFrame, QGroupBox, QMessageBox, QScrollArea, QDialog, QSizePolicy,
    QTableWidget, QTableWidgetItem, QHeaderView, QCompleter, QFileDialog,
    QTextEdit, QAbstractItemView, QSpinBox, QTabWidget, QRadioButton, QButtonGroup,
    QCheckBox, QMenu
)
from PyQt6.QtCore import (
    Qt, QTimer, pyqtSignal, QThread, QSize, QRect, QStringListModel, QPoint
)
from PyQt6.QtGui import QPixmap, QImage, QKeySequence, QShortcut, QColor, QAction

from igdb_api import igdb_client
from mobygames_api import moby_client
from retroachievements_api import ra_client
from config import API_CONFIG, save_settings, load_settings
from cache_manager import CacheManager


DARK_STYLESHEET = """
QMainWindow, QDialog { background-color: #1a1a2e; }
QWidget { background-color: #1a1a2e; color: #eaeaea; font-family: 'Segoe UI', Arial, sans-serif; font-size: 13px; }
QGroupBox { border: 2px solid #4a4a6a; border-radius: 8px; margin-top: 10px; padding-top: 8px; font-weight: bold; color: #00d4ff; }
QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 6px; }
QLabel { color: #b8b8d1; font-size: 12px; }
QLabel#titleLabel { color: #00d4ff; font-size: 22px; font-weight: bold; }
QLabel#resultLabel { color: #eaeaea; font-size: 12px; padding: 5px 8px; background-color: #252540; border-radius: 5px; border: 1px solid #3a3a5a; min-height: 16px; }
QLabel#coverLabel { background-color: #252540; border: 2px solid #3a3a5a; border-radius: 8px; }
QLabel#coverLabel:hover { border-color: #00d4ff; }
QLabel#screenshotThumb { background-color: #252540; border: 2px solid #3a3a5a; border-radius: 4px; }
QLabel#screenshotThumb:hover { border-color: #00d4ff; }
QLineEdit, QSpinBox { background-color: #252540; border: 2px solid #3a3a5a; border-radius: 5px; padding: 6px 10px; color: #eaeaea; font-size: 13px; }
QLineEdit:focus, QSpinBox:focus { border-color: #00d4ff; }
QTextEdit { background-color: #252540; border: 2px solid #3a3a5a; border-radius: 5px; padding: 5px; color: #eaeaea; font-size: 12px; }
QComboBox { background-color: #252540; border: 2px solid #3a3a5a; border-radius: 5px; padding: 6px 10px; color: #eaeaea; font-size: 13px; }
QComboBox:focus { border-color: #00d4ff; }
QComboBox:editable { background-color: #252540; }
QComboBox::drop-down { border: none; width: 25px; }
QComboBox::down-arrow { border-left: 4px solid transparent; border-right: 4px solid transparent; border-top: 6px solid #00d4ff; margin-right: 8px; }
QComboBox QAbstractItemView { background-color: #252540; border: 2px solid #00d4ff; selection-background-color: #3a3a6a; color: #eaeaea; }
QListWidget { background-color: #252540; border: 2px solid #3a3a5a; border-radius: 5px; padding: 2px; color: #eaeaea; outline: none; }
QListWidget::item { padding: 3px 8px; border-radius: 3px; margin: 1px 0; }
QListWidget::item:hover { background-color: #3a3a6a; }
QListWidget::item:selected { background-color: #00d4ff; color: #1a1a2e; }
QTableWidget { background-color: #252540; border: 2px solid #3a3a5a; gridline-color: #3a3a5a; color: #eaeaea; selection-background-color: #00d4ff; selection-color: #1a1a2e; }
QTableWidget::item { padding: 5px; }
QHeaderView::section { background-color: #3a3a5a; color: #00d4ff; padding: 8px; border: 1px solid #4a4a6a; font-weight: bold; }
QPushButton { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #00d4ff, stop:1 #0099cc); border: none; border-radius: 6px; padding: 8px 20px; color: #1a1a2e; font-size: 13px; font-weight: bold; min-width: 80px; }
QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #33e0ff, stop:1 #00b8e6); }
QPushButton:pressed { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #0099cc, stop:1 #007399); }
QPushButton:disabled { background: #4a4a6a; color: #8a8a9a; }
QPushButton#navButton { min-width: 30px; max-width: 30px; padding: 5px; font-size: 16px; }
QPushButton#refreshButton { min-width: 150px; padding: 8px; background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #aa4444, stop:1 #882222); font-weight: bold; border: 1px solid #ff5555; }
QPushButton#refreshButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #cc5555, stop:1 #aa3333); }

/* Навигация страниц */
QPushButton#pageNavButton { 
    min-width: 100px; 
    padding: 10px 15px; 
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #6644aa, stop:1 #5533aa);
    font-size: 14px;
    font-weight: bold;
    border: 1px solid #7755bb;
}
QPushButton#pageNavButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #7755bb, stop:1 #6644bb); }

/* Кнопка Очистить и Поиска мелкая */
QPushButton#iconButton {
    min-width: 40px; 
    max-width: 40px; 
    padding: 5px;
    background: #444466;
    border: 1px solid #666688;
    color: #ffcccc;
}
QPushButton#iconButton:hover { background: #555577; border-color: #777799; color: #ffffff; }

/* === КНОПКИ ИЗОБРАЖЕНИЙ (Обложка/Скриншоты) ВО "ВСЕ ИГРЫ" === */
QPushButton#imgBtnCover {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #ff8800, stop:1 #cc6600); /* Оранжевый */
    border: 1px solid #ffaa33;
    font-weight: bold;
    font-size: 13px;
    min-width: 120px;
}
QPushButton#imgBtnCover:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #ff9933, stop:1 #dd7700); }
QPushButton#imgBtnCover:disabled { background: #4a4a6a; border: none; color: #8a8a9a; }

QPushButton#imgBtnSS {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #9933ff, stop:1 #7722cc); /* Фиолетовый */
    border: 1px solid #aa55ff;
    font-weight: bold;
    font-size: 13px;
    min-width: 120px;
}
QPushButton#imgBtnSS:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #aa55ff, stop:1 #8833dd); }
QPushButton#imgBtnSS:disabled { background: #4a4a6a; border: none; color: #8a8a9a; }

QPushButton#closeButton { background: #ff4466; min-width: 40px; max-width: 40px; padding: 8px; font-size: 18px; }
QPushButton#closeButton:hover { background: #ff6688; }
QPushButton#allGamesButton { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #9966ff, stop:1 #6633cc); min-width: 200px; }
QFrame#separator { background-color: #3a3a5a; max-height: 2px; }
QScrollBar:vertical { background-color: #252540; width: 8px; border-radius: 4px; }
QScrollBar::handle:vertical { background-color: #4a4a6a; border-radius: 4px; min-height: 25px; }
QScrollBar::handle:vertical:hover { background-color: #00d4ff; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
QScrollBar:horizontal { background-color: #252540; height: 8px; border-radius: 4px; }
QScrollBar::handle:horizontal { background-color: #4a4a6a; border-radius: 4px; min-width: 25px; }
QScrollBar::handle:horizontal:hover { background-color: #00d4ff; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0px; }
QTabWidget::pane { border: 2px solid #4a4a6a; border-radius: 5px; background-color: #1a1a2e; }
QTabBar::tab { background-color: #252540; color: #b8b8d1; padding: 10px 20px; margin-right: 2px; border-top-left-radius: 5px; border-top-right-radius: 5px; }
QTabBar::tab:selected { background-color: #00d4ff; color: #1a1a2e; font-weight: bold; }
QTabBar::tab:hover:!selected { background-color: #3a3a6a; }
QRadioButton { color: #b8b8d1; spacing: 5px; }
QRadioButton::indicator { width: 13px; height: 13px; border-radius: 7px; border: 2px solid #3a3a5a; background-color: #252540; }
QRadioButton::indicator:checked { background-color: #00d4ff; border-color: #00d4ff; }

/* === НОВЫЕ СТИЛИ ДЛЯ ЧЕКБОКСОВ === */
QCheckBox { 
    spacing: 10px; 
    color: #eaeaea; 
    font-size: 13px;
    font-weight: bold;
}
QCheckBox::indicator { 
    width: 18px; 
    height: 18px; 
    border-radius: 4px; 
    border: 2px solid #4a4a6a; 
    background-color: #252540; 
}
QCheckBox::indicator:hover { 
    border-color: #00d4ff; 
}
QCheckBox::indicator:checked { 
    background-color: #00cc66; /* Зеленая заливка */
    border-color: #00cc66; 
    image: none; /* Убираем стандартную галочку, чтобы был просто квадрат */
}
/* Рамка для каждого пункта настроек */
QFrame#settingItem { 
    background-color: #202035; 
    border: 1px solid #3a3a5a; 
    border-radius: 6px; 
    padding: 2px;
}
QFrame#settingItem:hover {
    border-color: #00d4ff;
    background-color: #252540;
}

/* Контекстное меню */
QMenu {
    background-color: #252540;
    color: #eaeaea;
    border: 1px solid #4a4a6a;
    border-radius: 5px;
    padding: 5px;
}
QMenu::item {
    padding: 5px 20px 5px 10px;
    border-radius: 3px;
}
QMenu::item:selected {
    background-color: #00d4ff;
    color: #1a1a2e;
}
"""

# === CUSTOM CLASSES MOVED UP FOR VISIBILITY ===

class ClickableLabel(QLabel):
    clicked = pyqtSignal(int)
    def __init__(self, idx=0, parent=None): super().__init__(parent); self.idx = idx; self.setCursor(Qt.CursorShape.PointingHandCursor)
    def mousePressEvent(self, e):
        if e.button() == Qt.MouseButton.LeftButton: self.clicked.emit(self.idx)

class ContextMenuLabel(ClickableLabel):
    """Clickable Label with Context Menu for Save/Show in Folder"""
    def __init__(self, idx=0, parent=None):
        super().__init__(idx, parent)
        self.current_image_path = None

    def set_image_path(self, path):
        self.current_image_path = path

    def contextMenuEvent(self, event):
        if not self.current_image_path or not os.path.exists(self.current_image_path):
            return
        
        menu = QMenu(self)
        save_action = menu.addAction("Сохранить изображение...")
        show_action = menu.addAction("Показать в папке")
        
        action = menu.exec(self.mapToGlobal(event.pos()))
        
        if action == save_action:
            self.save_image()
        elif action == show_action:
            self.show_in_folder()

    def save_image(self):
        if not self.current_image_path: return
        file_ext = os.path.splitext(self.current_image_path)[1]
        save_path, _ = QFileDialog.getSaveFileName(self, "Сохранить изображение", "image" + file_ext, f"Images (*{file_ext})")
        if save_path:
            try:
                shutil.copy2(self.current_image_path, save_path)
                QMessageBox.information(self, "Успех", "Изображение сохранено")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить: {e}")

    def show_in_folder(self):
        if not self.current_image_path: return
        path = os.path.abspath(self.current_image_path)
        try:
            if os.name == 'nt':
                subprocess.run(['explorer', '/select,', path])
            elif sys.platform == 'darwin':
                subprocess.run(['open', '-R', path])
            else:
                subprocess.run(['xdg-open', os.path.dirname(path)])
        except Exception as e:
            print(f"Error opening folder: {e}")


# === WORKER THREADS ===

class ImageLoader(QThread):
    finished = pyqtSignal(QPixmap, str)
    error = pyqtSignal(str)
    
    def __init__(self, url: str, url_type: str = "cover", save_to_cache: bool = True):
        super().__init__()
        self.url = url
        self.url_type = url_type
        self.save_to_cache = save_to_cache
    
    def run(self):
        if not self.url:
            self.error.emit("Пустой URL")
            return

        # 1. Проверяем кэш
        local_path = CacheManager.get_image_path(self.url)
        
        if os.path.exists(local_path):
            image = QImage()
            if image.load(local_path):
                self.finished.emit(QPixmap.fromImage(image), self.url_type)
                return
        
        # 2. Если нет в кэше, качаем
        try:
            response = requests.get(self.url, timeout=10)
            response.raise_for_status()
            
            # Сохраняем в кэш, только если разрешено
            if self.save_to_cache:
                CacheManager.save_image(self.url, response.content)
            
            # Создаем изображение
            image = QImage()
            image.loadFromData(response.content)
            self.finished.emit(QPixmap.fromImage(image), self.url_type)
        except Exception as e:
            self.error.emit(str(e))


class BulkImageCacher(QThread):
    """Downloads images for all games in the list based on settings"""
    def __init__(self, games, cache_covers=False, cache_ss=False):
        super().__init__()
        self.games = games
        self.cache_covers = cache_covers
        self.cache_ss = cache_ss
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self):
        if not self.games: return
        
        # To avoid blocking network too much, we process sequentially
        for game in self.games:
            if self._stop: break
            
            did_download = False
            
            # Cache Cover
            if self.cache_covers:
                cover_url = game.get("cover_url_large")
                if cover_url and not CacheManager.has_image(cover_url):
                    try:
                        r = requests.get(cover_url, timeout=5)
                        if r.status_code == 200:
                            CacheManager.save_image(cover_url, r.content)
                            did_download = True
                    except: pass
                    
            # Cache Screenshots
            if self.cache_ss:
                ss_list = game.get("screenshot_urls", [])
                for ss in ss_list:
                    if self._stop: break
                    url = ss.get("medium") # Cache medium for list views/fast access
                    if url and not CacheManager.has_image(url):
                         try:
                            r = requests.get(url, timeout=5)
                            if r.status_code == 200:
                                CacheManager.save_image(url, r.content)
                                did_download = True
                         except: pass
            
            # FIX: Only sleep if we actually downloaded something to be polite.
            # If files exist, run as fast as possible.
            if did_download:
                time.sleep(0.05)


class RAThread(QThread):
    finished = pyqtSignal(dict)
    
    def __init__(self, game_name, platform_id):
        super().__init__()
        self.game_name = game_name
        self.platform_id = platform_id
        
    def run(self):
        try:
            cid = ra_client.get_console_id(int(self.platform_id)) if self.platform_id else None
            if not cid:
                self.finished.emit({})
                return
            
            gid = ra_client.get_game_id(cid, self.game_name)
            if gid:
                info = ra_client.get_game_info(gid)
                self.finished.emit(info)
            else:
                self.finished.emit({})
        except Exception as e:
            print(f"RA Error: {e}")
            self.finished.emit({})

class RAGameChecker(QThread):
    progress = pyqtSignal(int, str, str) # row, text, tooltip
    list_loaded = pyqtSignal(list)
    
    def __init__(self, console_id, targets, game_list_cache=None):
        super().__init__()
        self.console_id = console_id
        self.targets = targets # [(row, name), ...]
        self.game_list_cache = game_list_cache
        self._stop = False
        
    def stop(self): self._stop = True
    
    def run(self):
        if not self.console_id: return
        import time
        
        # 1. Get console game list
        if not self.game_list_cache:
            self.game_list_cache = ra_client.get_game_list_raw(self.console_id)
            if not self.game_list_cache: return
            self.list_loaded.emit(self.game_list_cache)

        # 2. Iterate targets (one page)
        for row, name in self.targets:
            if self._stop: break
            
            target_norm = ra_client._normalize(name)
            found_id = None
            found_title = ""
            
            # Fast scan in list
            for g in self.game_list_cache:
                if ra_client._normalize(g.get("Title", "")) == target_norm:
                    found_id = g.get("ID"); found_title = g.get("Title", ""); break
            
            if not found_id:
                for g in self.game_list_cache:
                    gn = ra_client._normalize(g.get("Title", ""))
                    if target_norm in gn or gn in target_norm:
                        found_id = g.get("ID"); found_title = g.get("Title", ""); break
            
            if found_id:
                info = ra_client.get_game_info(found_id)
                num = info.get("NumAchievements", 0)
                if num == 0 and "Achievements" in info:
                    num = len(info.get("Achievements", {}))
                self.progress.emit(row, f"{num} 🏆", f"{found_title} (ID: {found_id})")
            else:
                self.progress.emit(row, "—", "Не найдено в RetroAchievements")
            
            time.sleep(0.1)


class AllGamesLoader(QThread):
    finished = pyqtSignal(list)
    progress = pyqtSignal(int)
    error = pyqtSignal(str)
    
    def __init__(self, platform_id: int, source: str = "igdb"):
        super().__init__()
        self.platform_id = platform_id
        self.source = source
        self._stop = False
    
    def stop(self): self._stop = True
    
    def run(self):
        try:
            all_games = []
            offset = 0
            while not self._stop:
                if self.source == "igdb":
                    games = igdb_client.get_all_games_for_platform(self.platform_id, offset, 500)
                elif self.source == "moby":
                    # Moby doesn't support easy "fetch all" paging with high limits,
                    # but we'll try to call the client method.
                    # WARNING: Moby Platform IDs are different from IGDB. 
                    # If this is called with IGDB ID, it might fail or return empty.
                    games = moby_client.get_all_games_for_platform(self.platform_id, offset, 100)
                else:
                    games = []

                if not games: break
                all_games.extend(games)
                self.progress.emit(len(all_games))
                
                # Pagination logic
                if self.source == "igdb":
                    if len(games) < 500: break
                    offset += 500
                elif self.source == "moby":
                    if len(games) < 100: break
                    offset += 100
                    time.sleep(1.0) # Throttle Moby

            self.finished.emit(all_games)
        except Exception as e:
            self.error.emit(str(e))


class FranchiseWorker(QThread):
    """Fetches games for a franchise"""
    finished = pyqtSignal(list)
    error = pyqtSignal(str)
    
    def __init__(self, franchise_id, query=None, mode="id"):
        super().__init__()
        self.franchise_id = franchise_id
        self.query = query
        self.mode = mode
        
    def run(self):
        try:
            if self.mode == "search":
                # Search for franchises
                results = igdb_client.search_franchises(self.query)
                self.finished.emit(results)
            else:
                # Get games for franchise ID
                games = igdb_client.get_franchise_games(self.franchise_id)
                self.finished.emit(games)
        except Exception as e:
            self.error.emit(str(e))


class SearchWorker(QThread):
    finished = pyqtSignal(list)
    error = pyqtSignal(str)
    
    def __init__(self, q, pid, rid=None, source="igdb"): 
        super().__init__()
        self.q, self.pid, self.rid = q, pid, rid
        self.source = source
        
    def run(self):
        try: 
            if self.source == "igdb":
                self.finished.emit(igdb_client.search_games(self.q, self.pid, self.rid))
            elif self.source == "moby":
                # MobyGames search logic
                self.finished.emit(moby_client.search_games(self.q, self.pid))
        except Exception as e: 
            self.error.emit(str(e))


class DetailsWorker(QThread):
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)
    
    def __init__(self, gid, pid, rid=None, source="igdb"): 
        super().__init__()
        self.gid, self.pid, self.rid = gid, pid, rid
        self.source = source
        
    def run(self):
        cache_path = os.path.join(CacheManager.GAMES_DIR, f"{self.gid}_{self.pid}_{self.source}.json")
        
        cached_data = CacheManager.load_json(cache_path)
        # --- FIX: Check if cached data is stale (missing new fields like franchise_data) ---
        is_stale = False
        if cached_data:
            if self.source == "igdb":
                # Check for one of the new fields added recently (e.g. expansions/DLC updated logic)
                if "franchise_data" not in cached_data:
                    is_stale = True
        
        if cached_data and not is_stale:
            self.finished.emit(cached_data)
            return

        try:
            d = None
            if self.source == "igdb":
                d = igdb_client.get_game_details(self.gid, self.rid, self.pid)
            elif self.source == "moby":
                d = moby_client.get_game_details(self.gid)
                
            if d: 
                # Сохраняем в кэш только если разрешено
                if API_CONFIG.get("CACHE_SEARCH_INFO", True):
                    CacheManager.save_json(cache_path, d)
                self.finished.emit(d)
            else: 
                self.error.emit("Не найдено")
        except Exception as e: 
            self.error.emit(str(e))


class ImageViewer(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)
        self.pixmap = None
        self.active_loader = None
        self.current_url = None # Store URL to resolve path
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        
        top = QHBoxLayout()
        top.addStretch()
        close_btn = QPushButton("✕")
        close_btn.setObjectName("closeButton")
        close_btn.clicked.connect(self.close)
        top.addWidget(close_btn)
        layout.addLayout(top)
        
        # FIX: Replace QLabel with ContextMenuLabel for full screen
        self.image_label = ContextMenuLabel(parent=self)
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.image_label.setStyleSheet("background-color: transparent;")
        self.image_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        layout.addWidget(self.image_label, 1)
        
        QShortcut(QKeySequence(Qt.Key.Key_Escape), self, self.close)
    
    def show_image(self, url: str):
        self.current_url = url
        self.image_label.setText("Загрузка...")
        if self.active_loader and self.active_loader.isRunning():
            self.active_loader.terminate()
            self.active_loader.wait()
        
        self.active_loader = ImageLoader(url, "viewer", save_to_cache=True)
        self.active_loader.finished.connect(self.on_loaded)
        self.active_loader.start()
    
    def closeEvent(self, e):
        if self.active_loader and self.active_loader.isRunning():
            self.active_loader.terminate()
            self.active_loader.wait()
        super().closeEvent(e)
    
    def show_pixmap(self, pixmap: QPixmap):
        self.pixmap = pixmap
        # If we just show pixmap directly (like from main window), we might not know the path
        # But we can try to retrieve it from parent if available, or just leave it.
        # For now, updates image.
        self.update_image()
    
    def on_loaded(self, pixmap: QPixmap, _):
        self.pixmap = pixmap
        
        # Set path for context menu
        if self.current_url:
            path = CacheManager.get_image_path(self.current_url)
            self.image_label.set_image_path(path)
            
        self.update_image()
    
    def update_image(self):
        if self.pixmap and not self.pixmap.isNull():
            scaled = self.pixmap.scaled(self.image_label.size(), Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            self.image_label.setPixmap(scaled)
    
    def showEvent(self, e):
        super().showEvent(e)
        if self.parent():
            self.setGeometry(self.parent().geometry())
    
    def paintEvent(self, e):
        from PyQt6.QtGui import QPainter
        p = QPainter(self)
        p.fillRect(self.rect(), QColor(10, 10, 20, 230))


class ScreenshotViewer(ImageViewer):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.screenshots = []
        self.current_index = 0
        self.add_navigation()
    
    def add_navigation(self):
        layout = self.layout()
        nav = QHBoxLayout()
        self.prev_btn = QPushButton("◀")
        self.prev_btn.setObjectName("navButton")
        self.prev_btn.clicked.connect(self.show_prev)
        self.prev_btn.setFixedHeight(100)
        
        self.next_btn = QPushButton("▶")
        self.next_btn.setObjectName("navButton")
        self.next_btn.clicked.connect(self.show_next)
        self.next_btn.setFixedHeight(100)
        
        self.counter = QLabel("1/1")
        self.counter.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.counter.setStyleSheet("color: white; font-size: 14px; background: transparent;")
        
        # Replace image label layout
        img_layout = QHBoxLayout()
        img_layout.addWidget(self.prev_btn)
        img_layout.addWidget(self.image_label, 1)
        img_layout.addWidget(self.next_btn)
        
        # Insert before stretch
        layout.insertLayout(1, img_layout)
        layout.addWidget(self.counter)
        
        QShortcut(QKeySequence(Qt.Key.Key_Left), self, self.show_prev)
        QShortcut(QKeySequence(Qt.Key.Key_Right), self, self.show_next)
    
    def set_screenshots(self, ss: list, idx: int = 0):
        self.screenshots = ss
        self.current_index = idx
        self.load_current()
    
    def load_current(self):
        if self.screenshots and self.current_index < len(self.screenshots):
            self.show_image(self.screenshots[self.current_index].get("full", ""))
        self.counter.setText(f"{self.current_index + 1} / {len(self.screenshots)}")
        self.prev_btn.setEnabled(self.current_index > 0)
        self.next_btn.setEnabled(self.current_index < len(self.screenshots) - 1)
    
    def show_prev(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.load_current()
    
    def show_next(self):
        if self.current_index < len(self.screenshots) - 1:
            self.current_index += 1
            self.load_current()


class CopyableTableWidget(QTableWidget):
    def keyPressEvent(self, event):
        if event.matches(QKeySequence.StandardKey.Copy):
            selected = self.selectedRanges()
            if not selected: return
            s = ""
            for r in range(selected[0].topRow(), selected[0].bottomRow() + 1):
                for c in range(selected[0].leftColumn(), selected[0].rightColumn() + 1):
                    try: s += str(self.item(r, c).text()) + "\t"
                    except AttributeError: s += "\t"
                s = s[:-1] + "\n"
            QApplication.clipboard().setText(s)
        else:
            super().keyPressEvent(event)


class AllGamesDialog(QDialog):
    game_selected = pyqtSignal(str, int)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Все игры платформы")
        self.setMinimumSize(1200, 700)
        self.resize(1400, 800)
        self.games_data = []
        self.filtered_data = []
        self.current_page = 0
        self.games_per_page = 50
        self.loader, self.ra_worker, self.bulk_cacher = None, None, None
        self.ra_console_id = None
        self.ra_game_list_cache = None
        self.current_platform_name = ""
        self.current_platform_id = 0
        self.current_source = "igdb" 
        self.image_viewer, self.ss_viewer = None, None
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Header
        header = QHBoxLayout()
        self.title_label = QLabel("📋 Все игры платформы")
        self.title_label.setStyleSheet("color: #00d4ff; font-size: 18px; font-weight: bold;")
        header.addWidget(self.title_label)
        header.addStretch()
        self.status_label = QLabel("Загрузка...")
        header.addWidget(self.status_label)
        layout.addLayout(header)
        
        # --- FILTERS ---
        filters_box = QGroupBox("Фильтры")
        filters_layout = QHBoxLayout(filters_box)
        filters_layout.setSpacing(15)
        
        # Genres
        filters_layout.addWidget(QLabel("Жанр:"))
        self.filter_genre = QComboBox()
        self.filter_genre.addItem("Все")
        self.filter_genre.currentIndexChanged.connect(self.apply_filters)
        filters_layout.addWidget(self.filter_genre)
        
        # RA
        filters_layout.addWidget(QLabel("RetroAchievements:"))
        self.filter_ra = QComboBox()
        self.filter_ra.addItems(["Все", "Есть достижения", "Нет достижений"])
        self.filter_ra.currentIndexChanged.connect(self.apply_filters)
        filters_layout.addWidget(self.filter_ra)
        
        # Players
        filters_layout.addWidget(QLabel("Игроки:"))
        self.filter_players = QComboBox()
        self.filter_players.addItems(["Все", "1", "2", "3", "4+"])
        self.filter_players.currentIndexChanged.connect(self.apply_filters)
        filters_layout.addWidget(self.filter_players)
        
        # Devs
        filters_layout.addWidget(QLabel("Разработчик:"))
        self.filter_dev = QComboBox()
        self.filter_dev.addItem("Все")
        self.filter_dev.setFixedWidth(150)
        self.filter_dev.currentIndexChanged.connect(self.apply_filters)
        filters_layout.addWidget(self.filter_dev)
        
        # Pubs
        filters_layout.addWidget(QLabel("Издатель:"))
        self.filter_pub = QComboBox()
        self.filter_pub.addItem("Все")
        self.filter_pub.setFixedWidth(150)
        self.filter_pub.currentIndexChanged.connect(self.apply_filters)
        filters_layout.addWidget(self.filter_pub)
        
        # Franchise
        filters_layout.addWidget(QLabel("Франшиза:"))
        self.filter_franchise = QComboBox()
        self.filter_franchise.addItem("Все")
        self.filter_franchise.setFixedWidth(150)
        self.filter_franchise.currentIndexChanged.connect(self.apply_filters)
        filters_layout.addWidget(self.filter_franchise)
        
        layout.addWidget(filters_box)

        # Search
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 Поиск:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Введите название игры...")
        self.search_input.textChanged.connect(self.apply_filters)
        search_layout.addWidget(self.search_input, 1)
        layout.addLayout(search_layout)
        
        # Table
        self.table = CopyableTableWidget()
        headers = [
            "Название", "Издание", "Дата выхода", "Статус", "Жанры", 
            "RetroAchievements", "Регионы", "Русский", "Игроки", 
            "Разработчики", "Издатели", "Франшиза", "Осн. игра", "DLC", "Ремейки"
        ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.verticalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        header = self.table.horizontalHeader()
        for i in range(len(headers)): 
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        
        self.table.setColumnWidth(0, 200) # Name
        self.table.setColumnWidth(1, 100) # Edition
        self.table.setColumnWidth(2, 90)  # Date
        self.table.setColumnWidth(3, 80)  # Status
        self.table.setColumnWidth(4, 120) # Genres
        self.table.setColumnWidth(5, 100) # RA
        self.table.setColumnWidth(6, 90)  # Regions
        self.table.setColumnWidth(7, 70)  # RU
        self.table.setColumnWidth(8, 60)  # Players
        self.table.setColumnWidth(9, 120) # Devs
        self.table.setColumnWidth(10, 120) # Pubs
        
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        self.table.doubleClicked.connect(self.select_game)
        layout.addWidget(self.table)
        
        # Pagination
        page_layout = QHBoxLayout()
        self.prev_btn = QPushButton("◀ Предыдущая")
        self.prev_btn.setObjectName("pageNavButton")
        self.prev_btn.clicked.connect(self.prev_page)
        page_layout.addWidget(self.prev_btn)
        page_layout.addStretch()
        page_layout.addWidget(QLabel("Страница:"))
        self.page_spin = QSpinBox()
        self.page_spin.setMinimum(1)
        self.page_spin.valueChanged.connect(self.go_to_page)
        page_layout.addWidget(self.page_spin)
        self.page_label = QLabel("/ 1")
        page_layout.addWidget(self.page_label)
        page_layout.addStretch()
        self.next_btn = QPushButton("Следующая ▶")
        self.next_btn.setObjectName("pageNavButton")
        self.next_btn.clicked.connect(self.next_page)
        page_layout.addWidget(self.next_btn)
        layout.addLayout(page_layout)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        self.refresh_btn = QPushButton("🔄 Обновить кэш")
        self.refresh_btn.setStyleSheet("background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #aa4444, stop:1 #882222);")
        self.refresh_btn.clicked.connect(lambda: self.load_games(self.current_platform_name, self.current_platform_id, self.current_source, force=True))
        btn_layout.addWidget(self.refresh_btn)
        
        self.export_btn = QPushButton("📊 Экспорт в Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        btn_layout.addWidget(self.export_btn)
        
        btn_layout.addStretch()
        
        self.view_cover_btn = QPushButton("🖼 Обложка")
        self.view_cover_btn.setObjectName("imgBtnCover")
        self.view_cover_btn.setEnabled(False)
        self.view_cover_btn.clicked.connect(self.show_selected_cover)
        btn_layout.addWidget(self.view_cover_btn)
        
        self.view_ss_btn = QPushButton("📸 Скриншоты")
        self.view_ss_btn.setObjectName("imgBtnSS")
        self.view_ss_btn.setEnabled(False)
        self.view_ss_btn.clicked.connect(self.show_selected_ss)
        btn_layout.addWidget(self.view_ss_btn)
        
        btn_layout.addStretch()
        
        self.select_btn = QPushButton("✓ Выбрать игру")
        self.select_btn.clicked.connect(self.select_game)
        btn_layout.addWidget(self.select_btn)
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.close)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
    
    def load_games(self, platform_name: str, platform_id: int, source: str = "igdb", force: bool = False):
        self.current_platform_name = platform_name
        self.current_platform_id = platform_id
        self.current_source = source
        
        self.title_label.setText(f"📋 Все игры: {platform_name} ({source.upper()})")
        self.games_data = []
        self.filtered_data = []
        self.current_page = 0
        self.ra_console_id = ra_client.get_console_id(int(platform_id))
        self.ra_game_list_cache = None # Will be loaded by RAGameChecker
        
        # Reset filters UI
        self.block_filter_signals(True)
        self.filter_genre.setCurrentIndex(0)
        self.filter_ra.setCurrentIndex(0)
        self.filter_players.setCurrentIndex(0)
        self.filter_dev.clear(); self.filter_dev.addItem("Все")
        self.filter_pub.clear(); self.filter_pub.addItem("Все")
        self.filter_franchise.clear(); self.filter_franchise.addItem("Все")
        self.block_filter_signals(False)
        
        self.table.setRowCount(0)
        
        cache_path = CacheManager.get_list_path(platform_id, source)
        
        if not force and os.path.exists(cache_path):
            self.status_label.setText("Загрузка из кэша...")
            data = CacheManager.load_json(cache_path)
            if data:
                self.on_loaded(data, saved_to_cache=True)
                return

        self.status_label.setText("Загрузка с сайта (может занять время)...")
        if self.loader: self.loader.stop(); self.loader.wait()
        
        self.loader = AllGamesLoader(platform_id, source)
        self.loader.finished.connect(lambda d: self.on_loaded(d, saved_to_cache=False)) 
        self.loader.progress.connect(lambda c: self.status_label.setText(f"Загружено: {c}..."))
        self.loader.error.connect(lambda e: self.status_label.setText(f"Ошибка: {e}"))
        self.loader.start()
    
    def block_filter_signals(self, block):
        self.filter_genre.blockSignals(block)
        self.filter_ra.blockSignals(block)
        self.filter_players.blockSignals(block)
        self.filter_dev.blockSignals(block)
        self.filter_pub.blockSignals(block)
        self.filter_franchise.blockSignals(block)

    def on_loaded(self, games, saved_to_cache=False):
        self.games_data = sorted(games, key=lambda x: x.get("name", "").lower())
        
        # Populate Filters
        self.populate_filters()
        
        if not saved_to_cache and API_CONFIG.get("CACHE_ALL_INFO", True):
            CacheManager.save_json(CacheManager.get_list_path(self.current_platform_id, self.current_source), self.games_data)
            self.status_label.setText(f"Всего: {len(self.games_data)} (Сохранено в кэш)")
        else:
            self.status_label.setText(f"Всего: {len(self.games_data)}")
            
        self.filtered_data = self.games_data[:]
        self.update_table()
        
        cache_covers = API_CONFIG.get("CACHE_ALL_COVER", False)
        cache_ss = API_CONFIG.get("CACHE_ALL_SS", False)
        
        if cache_covers or cache_ss:
            if self.bulk_cacher: self.bulk_cacher.stop(); self.bulk_cacher.wait()
            self.status_label.setText(f"Всего: {len(self.games_data)} (Кэширование изображений...)")
            self.bulk_cacher = BulkImageCacher(self.games_data, cache_covers, cache_ss)
            # FIX: Connect finished signal
            self.bulk_cacher.finished.connect(self.on_caching_finished)
            self.bulk_cacher.start()

    def on_caching_finished(self):
        self.status_label.setText(f"Всего: {len(self.games_data)} (Кэширование завершено)")
    
    def populate_filters(self):
        self.block_filter_signals(True)
        
        genres = set()
        devs = set()
        pubs = set()
        franchises = set()
        
        for g in self.games_data:
            # Genres string split
            g_str = g.get("mixed_genres", "")
            if not g_str: g_str = igdb_client.format_genres(g.get("genres"))
            if g_str:
                for x in g_str.split("; "): genres.add(x)
            
            # Devs
            d_list = g.get("developers_list", [])
            for d in d_list: devs.add(d)
            
            # Pubs
            p_list = g.get("publishers_list", [])
            for p in p_list: pubs.add(p)
            
            # Franchise
            f_str = g.get("franchise_str", "")
            if f_str:
                for f in f_str.split("; "): franchises.add(f)
        
        # Populate Combos
        self.filter_genre.clear(); self.filter_genre.addItem("Все"); self.filter_genre.addItems(sorted(list(genres)))
        self.filter_dev.clear(); self.filter_dev.addItem("Все"); self.filter_dev.addItems(sorted(list(devs)))
        self.filter_pub.clear(); self.filter_pub.addItem("Все"); self.filter_pub.addItems(sorted(list(pubs)))
        self.filter_franchise.clear(); self.filter_franchise.addItem("Все"); self.filter_franchise.addItems(sorted(list(franchises)))
        
        self.block_filter_signals(False)

    def apply_filters(self):
        txt = self.search_input.text().lower()
        f_genre = self.filter_genre.currentText()
        f_ra = self.filter_ra.currentText()
        f_players = self.filter_players.currentText()
        f_dev = self.filter_dev.currentText()
        f_pub = self.filter_pub.currentText()
        f_franchise = self.filter_franchise.currentText()
        
        filtered = []
        
        # Pre-load RA names for fast filtering if needed
        ra_names_set = set()
        if f_ra != "Все" and self.ra_game_list_cache:
             for item in self.ra_game_list_cache:
                 ra_names_set.add(ra_client._normalize(item.get("Title", "")))

        for g in self.games_data:
            # Text Search
            if txt and txt not in g.get("name", "").lower(): continue
            
            # Genre
            g_str = g.get("mixed_genres", "")
            if not g_str: g_str = igdb_client.format_genres(g.get("genres"))
            if f_genre != "Все" and f_genre not in g_str: continue
            
            # Players
            players = g.get("max_local_players", 1)
            if f_players != "Все":
                if f_players == "4+" and players < 4: continue
                elif f_players != "4+" and players != int(f_players): continue
            
            # Dev
            if f_dev != "Все" and f_dev not in g.get("developers_list", []): continue
            
            # Pub
            if f_pub != "Все" and f_pub not in g.get("publishers_list", []): continue
            
            # Franchise
            if f_franchise != "Все" and f_franchise not in g.get("franchise_str", ""): continue
            
            # RA Filter
            if f_ra != "Все":
                # Check if game is in RA list
                norm_name = ra_client._normalize(g.get("name", ""))
                has_ra = norm_name in ra_names_set
                # Also check partial match if direct fail? (simplified for speed)
                if not has_ra:
                     # Fallback check
                     for rn in ra_names_set:
                         if norm_name in rn: 
                             has_ra = True; break
                
                if f_ra == "Есть достижения" and not has_ra: continue
                if f_ra == "Нет достижений" and has_ra: continue

            filtered.append(g)
            
        self.filtered_data = filtered
        self.current_page = 0
        self.update_table()
    
    def filter_games(self, text):
        # Redirect to main filter function
        self.apply_filters()
    
    def update_table(self):
        if self.ra_worker: self.ra_worker.stop(); self.ra_worker.wait()
        
        start = self.current_page * self.games_per_page
        end = min(start + self.games_per_page, len(self.filtered_data))
        data = self.filtered_data[start:end]
        self.table.setRowCount(len(data))
        ra_targets = []
        
        for row, g in enumerate(data):
            name = g.get("name", "")
            item = QTableWidgetItem(name)
            item.setData(Qt.ItemDataRole.UserRole, g.get("id")) # Store Game ID
            item.setData(Qt.ItemDataRole.UserRole + 1, g)       # Store full game data
            self.table.setItem(row, 0, item)
            
            self.table.setItem(row, 1, QTableWidgetItem(g.get("edition", "")))
            
            date_item = QTableWidgetItem(igdb_client.format_release_date(g.get("earliest_release_date")))
            date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 2, date_item)

            self.table.setItem(row, 3, QTableWidgetItem(g.get("status_display", "—")))

            # Genres
            genres_display = g.get("mixed_genres", "")
            if not genres_display: genres_display = igdb_client.format_genres(g.get("genres"))
            self.table.setItem(row, 4, QTableWidgetItem(genres_display))
            
            # RA Column (5)
            ra_item = QTableWidgetItem("...")
            ra_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 5, ra_item)
            ra_targets.append((row, name))
            
            # Regions Column (6)
            reg_list = g.get("all_regions", [])
            self.table.setItem(row, 6, QTableWidgetItem("; ".join(reg_list) if reg_list else "—"))

            # Russian
            self.table.setItem(row, 7, QTableWidgetItem(g.get("russian_support", "None")))
            
            players_item = QTableWidgetItem(str(g.get("max_local_players", 1)))
            players_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 8, players_item)
            
            self.table.setItem(row, 9, QTableWidgetItem(igdb_client.format_list(g.get("developers_list"))))
            self.table.setItem(row, 10, QTableWidgetItem(igdb_client.format_list(g.get("publishers_list"))))
            
            self.table.setItem(row, 11, QTableWidgetItem(g.get("franchise_str", "")))
            self.table.setItem(row, 12, QTableWidgetItem(g.get("parent_game_str", "")))
            self.table.setItem(row, 13, QTableWidgetItem(g.get("dlcs_str", "")))
            self.table.setItem(row, 14, QTableWidgetItem(g.get("remakes_remasters_str", "")))
        
        # Start RA checker
        if self.ra_console_id:
            self.ra_worker = RAGameChecker(self.ra_console_id, ra_targets, self.ra_game_list_cache)
            self.ra_worker.progress.connect(self.on_ra_progress)
            self.ra_worker.list_loaded.connect(self.on_ra_list_loaded)
            self.ra_worker.start()
            
        total = max(1, (len(self.filtered_data) + self.games_per_page - 1) // self.games_per_page)
        self.page_spin.setMaximum(total)
        self.page_spin.blockSignals(True); self.page_spin.setValue(self.current_page + 1); self.page_spin.blockSignals(False)
        self.page_label.setText(f"/ {total}")
        self.prev_btn.setEnabled(self.current_page > 0)
        self.next_btn.setEnabled(self.current_page < total - 1)

    def on_ra_progress(self, row, text, tooltip):
        if row < self.table.rowCount():
            item = QTableWidgetItem(text)
            item.setToolTip(tooltip)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if "🏆" in text: item.setForeground(QColor("#ffcc00"))
            self.table.setItem(row, 5, item)

    def on_ra_list_loaded(self, data):
        self.ra_game_list_cache = data
        # If RA filter is active, re-apply filters now that we have data
        if self.filter_ra.currentIndex() > 0:
            self.apply_filters()
    
    def on_selection_changed(self):
        row = self.table.currentRow()
        has_sel = row >= 0
        self.view_cover_btn.setEnabled(has_sel)
        self.view_ss_btn.setEnabled(has_sel)

    def get_selected_game_data(self):
        row = self.table.currentRow()
        if row >= 0:
            item = self.table.item(row, 0)
            if item:
                return item.data(Qt.ItemDataRole.UserRole + 1)
        return None

    def show_selected_cover(self):
        g = self.get_selected_game_data()
        if g and g.get("cover_url_large"):
            if not self.image_viewer: self.image_viewer = ImageViewer(self)
            self.image_viewer.show_image(g.get("cover_url_large"))
            self.image_viewer.show()
        else:
            QMessageBox.information(self, "Инфо", "Нет обложки")

    def show_selected_ss(self):
        g = self.get_selected_game_data()
        if g and g.get("screenshot_urls"):
            if not self.ss_viewer: self.ss_viewer = ScreenshotViewer(self)
            self.ss_viewer.set_screenshots(g.get("screenshot_urls"), 0)
            self.ss_viewer.show()
        else:
            QMessageBox.information(self, "Инфо", "Нет скриншотов")

    def prev_page(self):
        if self.current_page > 0: self.current_page -= 1; self.update_table()
    
    def next_page(self):
        total = (len(self.filtered_data) + self.games_per_page - 1) // self.games_per_page
        if self.current_page < total - 1: self.current_page += 1; self.update_table()
    
    def go_to_page(self, page):
        self.current_page = page - 1
        self.update_table()
    
    def select_game(self):
        row = self.table.currentRow()
        if row >= 0:
            item = self.table.item(row, 0)
            if item: self.game_selected.emit(item.text(), item.data(Qt.ItemDataRole.UserRole)); self.close()
    
    def export_to_excel(self):
        if not self.filtered_data: QMessageBox.warning(self, "Внимание", "Нет данных"); return
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить", "", "Excel (*.xlsx)")
        if not path: return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook(); ws = wb.active; ws.title = "Games"
            headers = ["Название", "Издание", "Дата выхода", "Статус", "Жанры", "Регион", "Русский", "Игроки", "Разработчики", "Издатели", "Описание"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(1, c, h); cell.fill = PatternFill("solid", "00D4FF"); cell.font = Font(bold=True)
            for r, g in enumerate(self.filtered_data, 2):
                ws.cell(r, 1, g.get("name", ""))
                ws.cell(r, 2, g.get("edition", ""))
                ws.cell(r, 3, igdb_client.format_release_date(g.get("earliest_release_date")))
                ws.cell(r, 4, g.get("status_display", ""))
                
                # Mixed genres
                g_str = g.get("mixed_genres", "")
                if not g_str: g_str = igdb_client.format_genres(g.get("genres"))
                ws.cell(r, 5, g_str)
                
                ws.cell(r, 6, "; ".join(g.get("all_regions", [])))
                ws.cell(r, 7, g.get("russian_support", ""))
                ws.cell(r, 8, g.get("max_local_players", 1))
                ws.cell(r, 9, igdb_client.format_list(g.get("developers_list")))
                ws.cell(r, 10, igdb_client.format_list(g.get("publishers_list")))
                ws.cell(r, 11, g.get("description", ""))
            wb.save(path)
            QMessageBox.information(self, "Успех", f"Сохранено: {path}")
        except Exception as e: QMessageBox.critical(self, "Ошибка", str(e))
    
    def closeEvent(self, e):
        if self.loader: self.loader.stop(); self.loader.wait()
        if self.ra_worker: self.ra_worker.stop(); self.ra_worker.wait()
        if self.bulk_cacher: self.bulk_cacher.stop(); self.bulk_cacher.wait()
        super().closeEvent(e)


class SettingsTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        # Scroll area for settings
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        content = QWidget()
        self.form_layout = QVBoxLayout(content)
        self.form_layout.setSpacing(20)
        
        self.inputs = {}
        self.checks = {}
        
        # IGDB
        self.add_group("IGDB", [
            ("Client ID", "IGDB_CLIENT_ID"),
            ("Client Secret", "IGDB_CLIENT_SECRET")
        ])
        
        # MobyGames
        self.add_group("MobyGames", [
            ("API Key", "MOBYGAMES_API_KEY")
        ])
        
        # RetroAchievements
        self.add_group("RetroAchievements", [
            ("User", "RETROACHIEVEMENTS_USER"),
            ("API Key", "RETROACHIEVEMENTS_API_KEY")
        ])
        
        # === Caching Settings Group ===
        cache_group = QGroupBox("Кэширование")
        cache_layout = QVBoxLayout(cache_group)
        cache_layout.setSpacing(10)
        
        # Search Cache
        cache_layout.addWidget(QLabel("<b>Поиск игры:</b>"))
        
        self.add_check(cache_layout, "CACHE_SEARCH_INFO", "Информация")
        self.add_check(cache_layout, "CACHE_SEARCH_COVER", "Обложка")
        self.add_check(cache_layout, "CACHE_SEARCH_SS", "Скриншоты")
        
        cache_layout.addSpacing(15)
        
        # All Games Cache
        cache_layout.addWidget(QLabel("<b>Все игры (список):</b>"))
        self.add_check(cache_layout, "CACHE_ALL_INFO", "Информация (JSON)")
        self.add_check(cache_layout, "CACHE_ALL_COVER", "Обложка")
        self.add_check(cache_layout, "CACHE_ALL_SS", "Скриншоты")
        
        self.form_layout.addWidget(cache_group)
        
        self.form_layout.addStretch()
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        # Save button
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        save_btn = QPushButton("💾 Сохранить настройки")
        save_btn.clicked.connect(self.save_data)
        save_btn.setMinimumHeight(40)
        btn_layout.addWidget(save_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
    
    def add_group(self, title, fields):
        group = QGroupBox(title)
        layout = QVBoxLayout(group)
        for label, key, *args in fields:
            is_pass = args[0] if args else False
            row = QHBoxLayout()
            lbl = QLabel(label + ":")
            lbl.setFixedWidth(120)
            inp = QLineEdit()
            if is_pass: inp.setEchoMode(QLineEdit.EchoMode.Password)
            self.inputs[key] = inp
            row.addWidget(lbl)
            row.addWidget(inp)
            layout.addLayout(row)
        self.form_layout.addWidget(group)
        
    def add_check(self, layout, key, title):
        # Create a frame for "edging"
        frame = QFrame()
        frame.setObjectName("settingItem")
        # Layout inside frame
        fl = QHBoxLayout(frame)
        fl.setContentsMargins(10, 8, 10, 8)
        
        chk = QCheckBox(title)
        self.checks[key] = chk
        fl.addWidget(chk)
        
        layout.addWidget(frame)
    
    def load_data(self):
        cfg = load_settings()
        for key, inp in self.inputs.items():
            inp.setText(cfg.get(key, ""))
        
        # Load checkboxes
        for key, chk in self.checks.items():
            chk.setChecked(cfg.get(key, False))
            
    def save_data(self):
        cfg = {}
        for key, inp in self.inputs.items():
            cfg[key] = inp.text().strip()
        
        for key, chk in self.checks.items():
            cfg[key] = chk.isChecked()
            
        save_settings(cfg)
        API_CONFIG.update(cfg)
        
        # Re-init clients
        igdb_client.client_id = cfg.get("IGDB_CLIENT_ID", "")
        igdb_client.client_secret = cfg.get("IGDB_CLIENT_SECRET", "")
        igdb_client.access_token = None
        moby_client.api_key = cfg.get("MOBYGAMES_API_KEY", "")
        
        QMessageBox.information(self, "Успех", "Настройки сохранены!")


class IconLoader(QThread):
    finished = pyqtSignal(QPixmap, object)
    def __init__(self, url, target):
        super().__init__()
        self.url = url
        self.target = target
    def run(self):
        try:
            r = requests.get(self.url, timeout=15)
            if r.status_code == 200:
                img = QImage(); img.loadFromData(r.content)
                self.finished.emit(QPixmap.fromImage(img), self.target)
        except: pass

class AchievementsDialog(QDialog):
    def __init__(self, parent=None, achievements=None, game_title=""):
        super().__init__(parent)
        self.setWindowTitle(f"🏆 Достижения: {game_title}")
        self.setMinimumSize(600, 500)
        self.setStyleSheet("background-color: #1a1a2e; color: #eaeaea;")
        self.loaders = [] 
        layout = QVBoxLayout(self)
        self.list_widget = QListWidget()
        self.list_widget.setStyleSheet("QListWidget { border: none; background: #1a1a2e; } QListWidget::item { border-bottom: 1px solid #4a4a6a; padding: 10px; }")
        layout.addWidget(self.list_widget)
        
        achs_list = []
        if isinstance(achievements, dict): achs_list = list(achievements.values())
        elif isinstance(achievements, list): achs_list = achievements
        
        try: achs_list.sort(key=lambda x: int(x.get("DisplayOrder", 0)))
        except: pass
        
        for a in achs_list:
            item = QListWidgetItem()
            w = QWidget(); l = QHBoxLayout(w); l.setContentsMargins(0,0,0,0); l.setSpacing(15)
            
            icon_lbl = QLabel(); icon_lbl.setFixedSize(48, 48); icon_lbl.setStyleSheet("background: #2a2a4a; border-radius: 4px;")
            l.addWidget(icon_lbl)
            
            title = a.get("Title", "???"); desc = a.get("Description", "")
            points = a.get("Points", "0")
            
            txt = QLabel(f"<b style='font-size:14px; color: #ffcc00;'>{title}</b> <span style='color: #888;'>({points} pts)</span><br>{desc}")
            txt.setTextFormat(Qt.TextFormat.RichText); txt.setWordWrap(True)
            l.addWidget(txt, 1)
            
            item.setSizeHint(QSize(500, 90)) # Increased height for full text
            self.list_widget.addItem(item); self.list_widget.setItemWidget(item, w)
            
            badge = a.get("BadgeName", "")
            if badge:
                url = f"https://media.retroachievements.org/Badge/{badge}.png"
                loader = IconLoader(url, icon_lbl)
                loader.finished.connect(self.on_icon_loaded)
                loader.start()
                self.loaders.append(loader)
            
        cls_btn = QPushButton("Закрыть"); cls_btn.clicked.connect(self.close)
        layout.addWidget(cls_btn)
        
    def on_icon_loaded(self, pix, lbl):
        if not pix.isNull(): lbl.setPixmap(pix.scaled(48, 48, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("STORM GAMES INFO")
        self.setMinimumSize(1300, 850) # Increased height
        self.resize(1400, 900)
        self.platform_id = None
        self.search_results, self.selected_game_id, self.current_game_details = [], None, None
        self.search_worker, self.details_worker, self.franchise_worker, self.image_loaders = None, None, None, []
        self.current_ss_page, self.ss_per_page = 0, 4
        self._updating = False
        self.search_timer = QTimer(); self.search_timer.setSingleShot(True); self.search_timer.timeout.connect(self.perform_search)
        self.ss_viewer, self.all_games_dialog, self.cover_viewer, self.cover_pixmap = None, None, None, None
        self.current_franchise_games = [] # Store franchise data for filtering
        self.navigation_stack = [] # History stack for Back button logic
        self.init_ui()
    
    def init_ui(self):
        central = QWidget(); self.setCentralWidget(central)
        main_layout = QVBoxLayout(central); main_layout.setSpacing(10); main_layout.setContentsMargins(15, 15, 15, 15)
        
        # Title
        title = QLabel("🎮 STORM GAMES INFO"); title.setObjectName("titleLabel"); title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title)
        sep = QFrame(); sep.setObjectName("separator"); sep.setFrameShape(QFrame.Shape.HLine); main_layout.addWidget(sep)
        
        # Tabs
        tabs = QTabWidget()
        main_layout.addWidget(tabs)
        
        # === SEARCH TAB ===
        search_tab = QWidget()
        tabs.addTab(search_tab, "🔍 Поиск игр")
        
        content = QHBoxLayout(search_tab); content.setSpacing(15); content.setContentsMargins(10, 10, 10, 10)
        
        # Left - Cover & Screenshots
        left = QWidget(); left.setFixedWidth(280); left_l = QVBoxLayout(left); left_l.setContentsMargins(0,0,0,0); left_l.setSpacing(10)
        cover_g = QGroupBox("🖼️ Обложка"); cover_l = QVBoxLayout(cover_g); cover_l.setContentsMargins(10,15,10,10)
        self.cover_label = ContextMenuLabel(); self.cover_label.setObjectName("coverLabel"); self.cover_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.cover_label.setMinimumHeight(450); self.cover_label.setMaximumHeight(500); self.cover_label.setText("Нет изображения")
        self.cover_label.clicked.connect(self.show_cover_fullscreen); cover_l.addWidget(self.cover_label); left_l.addWidget(cover_g)
        
        ss_g = QGroupBox("📸 Скриншоты"); ss_l = QVBoxLayout(ss_g); ss_l.setContentsMargins(10,15,10,10); ss_l.setSpacing(8)
        ss_cont = QWidget(); ss_grid = QHBoxLayout(ss_cont); ss_grid.setContentsMargins(0,0,0,0); ss_grid.setSpacing(5)
        self.ss_labels = []
        for i in range(4):
            lbl = ContextMenuLabel(i); lbl.setObjectName("screenshotThumb"); lbl.setFixedSize(60,45); lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setScaledContents(True); lbl.clicked.connect(self.on_ss_clicked); ss_grid.addWidget(lbl); self.ss_labels.append(lbl)
        ss_l.addWidget(ss_cont)
        nav = QHBoxLayout(); nav.setSpacing(10)
        self.prev_ss = QPushButton("◀"); self.prev_ss.setObjectName("navButton"); self.prev_ss.clicked.connect(self.prev_ss_page); self.prev_ss.setEnabled(False); nav.addWidget(self.prev_ss)
        nav.addStretch(); self.ss_counter = QLabel("0 / 0"); self.ss_counter.setAlignment(Qt.AlignmentFlag.AlignCenter); nav.addWidget(self.ss_counter); nav.addStretch()
        self.next_ss = QPushButton("▶"); self.next_ss.setObjectName("navButton"); self.next_ss.clicked.connect(self.next_ss_page); self.next_ss.setEnabled(False); nav.addWidget(self.next_ss)
        ss_l.addLayout(nav); left_l.addWidget(ss_g)
        
        # Middle - Search
        mid = QWidget(); mid_l = QVBoxLayout(mid); mid_l.setContentsMargins(0,0,0,0); mid_l.setSpacing(8)
        
        platforms = igdb_client.get_platforms_list()
        search_g = QGroupBox(f"🔍 Поиск игры"); search_l = QVBoxLayout(search_g); search_l.setSpacing(8)
        
        # Source Selection (IGDB vs MobyGames)
        source_layout = QHBoxLayout()
        source_layout.addWidget(QLabel("Источник:"))
        self.rb_igdb = QRadioButton("IGDB")
        self.rb_igdb.setChecked(True)
        self.rb_moby = QRadioButton("MobyGames")
        
        self.source_group = QButtonGroup(self)
        self.source_group.addButton(self.rb_igdb)
        self.source_group.addButton(self.rb_moby)
        self.source_group.buttonClicked.connect(self.on_source_changed)
        
        source_layout.addWidget(self.rb_igdb)
        source_layout.addWidget(self.rb_moby)
        source_layout.addStretch()
        search_l.addLayout(source_layout)

        pl = QHBoxLayout()
        pl_lbl = QLabel("Платформа:"); pl_lbl.setFixedWidth(95); pl.addWidget(pl_lbl)
        self.platform_combo = QComboBox(); self.platform_combo.setEditable(True)
        self.platform_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.platform_combo.addItems(platforms)
        comp = QCompleter(platforms); comp.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive); comp.setFilterMode(Qt.MatchFlag.MatchContains)
        self.platform_combo.setCompleter(comp); self.platform_combo.currentIndexChanged.connect(self.on_platform_change)
        pl.addWidget(self.platform_combo, 1); search_l.addLayout(pl)
        

        
        agl = QHBoxLayout(); agl.addStretch()
        self.all_games_btn = QPushButton("📋 Все игры платформы"); self.all_games_btn.setObjectName("allGamesButton")
        self.all_games_btn.clicked.connect(self.show_all_games); agl.addWidget(self.all_games_btn); agl.addStretch()
        search_l.addLayout(agl)
        
        gl = QHBoxLayout()
        gl_lbl = QLabel("Название:"); gl_lbl.setFixedWidth(95); gl.addWidget(gl_lbl)
        self.game_input = QLineEdit(); self.game_input.setPlaceholderText("Минимум 2 символа...")
        self.game_input.textChanged.connect(self.on_game_input); gl.addWidget(self.game_input, 1); search_l.addLayout(gl)
        
        # --- FRANCHISE INPUT ---
        fl = QHBoxLayout()
        fl_lbl = QLabel("Франшиза:"); fl_lbl.setFixedWidth(95); fl.addWidget(fl_lbl)
        self.franchise_input = QLineEdit(); self.franchise_input.setPlaceholderText("Поиск по франшизе...")
        self.franchise_btn = QPushButton("🔍"); self.franchise_btn.setObjectName("iconButton")
        self.franchise_btn.setToolTip("Найти франшизы")
        self.franchise_btn.clicked.connect(self.search_franchises_manual)
        fl.addWidget(self.franchise_input, 1); fl.addWidget(self.franchise_btn); search_l.addLayout(fl)

        # --- List Widget (was Suggestions) ---
        search_l.addWidget(QLabel("📋 Список:"))
        self.suggestions = QListWidget(); self.suggestions.setMinimumHeight(100); self.suggestions.itemClicked.connect(self.on_suggestion_click)
        self.suggestions.itemDoubleClicked.connect(self.on_suggestion_dblclick); search_l.addWidget(self.suggestions)
        
        # --- КНОПКИ ПОИСКА и ОЧИСТКИ ---
        bl = QHBoxLayout(); bl.addStretch()
        self.search_btn = QPushButton("🔍 Искать")
        self.search_btn.clicked.connect(self.fetch_details)
        self.search_btn.setEnabled(False)
        bl.addWidget(self.search_btn)
        
        self.clear_btn = QPushButton("✕")
        self.clear_btn.setObjectName("clearButton")
        self.clear_btn.setToolTip("Очистить все поля")
        self.clear_btn.clicked.connect(self.clear_all)
        bl.addWidget(self.clear_btn)
        
        bl.addStretch()
        search_l.addLayout(bl)
        mid_l.addWidget(search_g)
        
        # Right - Results
        right = QWidget(); right_l = QVBoxLayout(right); right_l.setContentsMargins(0,0,0,0); right_l.setSpacing(8)
        res_g = QGroupBox("📊 Информация об игре"); res_l = QVBoxLayout(res_g); res_l.setSpacing(6)
        
        # Basic Info
        self.name_res = self._row(res_l, "Название:")
        self.date_res = self._row(res_l, "Дата выхода:")
        self.status_res = self._row(res_l, "Статус:") # New
        self.genres_res = self._row(res_l, "Жанры:")
        self.regions_res = self._row(res_l, "Регионы:")
        self.ru_res = self._row(res_l, "Русский язык:") # New (Localization)
        
        self.players_res = self._row(res_l, "Игроки:")
        self.devs_res = self._row(res_l, "Разработчики:")
        self.pubs_res = self._row(res_l, "Издатели:")
        
        # Relations (New)
        # We enable OpenLinks so we can use HTML anchors for clicks
        self.franchise_res = self._row(res_l, "Франшиза:")
        self.franchise_res.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        self.franchise_res.linkActivated.connect(self.load_game_from_link)

        self.parent_res = self._row(res_l, "Основная игра:")
        self.parent_res.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        self.parent_res.linkActivated.connect(self.load_game_from_link)

        self.dlc_res = self._row(res_l, "DLC:")
        self.dlc_res.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        self.dlc_res.linkActivated.connect(self.load_game_from_link)

        self.remake_res = self._row(res_l, "Ремейки:")
        self.remake_res.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        self.remake_res.linkActivated.connect(self.load_game_from_link)

        # RetroAchievements Row
        ra_h = QHBoxLayout(); ra_h.setSpacing(8)
        ra_lbl = QLabel("RetroAchievements:"); ra_lbl.setFixedWidth(130) # Changed label and width
        self.ra_res = QLabel("—"); self.ra_res.setTextFormat(Qt.TextFormat.RichText); self.ra_res.setOpenExternalLinks(True)
        self.ra_res.setStyleSheet("font-weight: bold; color: #ffcc00;")
        self.ra_res.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        
        self.ra_list_btn = QPushButton("📜"); self.ra_list_btn.setFixedSize(30, 25); self.ra_list_btn.setToolTip("Список достижений")
        self.ra_list_btn.clicked.connect(self.show_ra_list); self.ra_list_btn.setEnabled(False)
        self.ra_list_btn.setStyleSheet("background: #2a2a4a; border: 1px solid #4a4a6a; border-radius: 4px;")
        
        ra_h.addWidget(ra_lbl); ra_h.addWidget(self.ra_res, 1); ra_h.addWidget(self.ra_list_btn)
        res_l.addLayout(ra_h)
        
        res_l.addWidget(QLabel("Описание:"))
        self.desc_text = QTextEdit(); self.desc_text.setReadOnly(True); self.desc_text.setMinimumHeight(80); res_l.addWidget(self.desc_text)
        
        # --- КНОПКА ОБНОВЛЕНИЯ В БЛОКЕ ИНФОРМАЦИИ ---
        res_l.addStretch() # Прижать к низу
        
        refresh_layout = QHBoxLayout()
        refresh_layout.addStretch()
        self.refresh_btn = QPushButton("🔄 Обновить данные (сбросить кэш)")
        self.refresh_btn.setObjectName("refreshButton")
        self.refresh_btn.setToolTip("Удалить сохраненные данные об этой игре и загрузить заново")
        self.refresh_btn.clicked.connect(self.force_refresh)
        refresh_layout.addWidget(self.refresh_btn)
        refresh_layout.addStretch()
        res_l.addLayout(refresh_layout)
        
        right_l.addWidget(res_g)
        
        content.addWidget(left); content.addWidget(mid, 50); content.addWidget(right, 50)
        
        # === SETTINGS TAB ===
        tabs.addTab(SettingsTab(self), "⚙️ Настройки")
        
        self.status = QLabel("Готово"); self.status.setAlignment(Qt.AlignmentFlag.AlignCenter); self.status.setStyleSheet("color: #6a6a8a; font-size: 11px;"); main_layout.addWidget(self.status)
    
    def _row(self, layout, text):
        r = QHBoxLayout(); r.setSpacing(8); lbl = QLabel(text); lbl.setFixedWidth(130); res = QLabel("—"); res.setObjectName("resultLabel"); res.setWordWrap(True)
        res.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        r.addWidget(lbl); r.addWidget(res, 1); layout.addLayout(r); return res
    
    def on_source_changed(self, _):
        self.clear_all()
        self.status.setText("Источник изменен")

    def on_platform_change(self, _):
        try:
            if self._updating: return
            self._updating = True; self.game_input.clear(); self._updating = False
            self.platform_id = igdb_client.get_platform_id(self.platform_combo.currentText())
            self.search_timer.stop()
            self.suggestions.clear(); self.search_results = []; self.selected_game_id = None; self.search_btn.setEnabled(False)
        except Exception as e: print(f"Error in on_platform_change: {e}")
    
    def on_region_change(self, _):
        try:
            if self._updating: return
            self.search_timer.stop()
            self.suggestions.clear(); self.search_results = []; self.selected_game_id = None; self.search_btn.setEnabled(False)
            text = self.game_input.text() if self.game_input else ""
            if len(text) >= 2: self.search_timer.start(300)
        except Exception as e: print(f"Error in on_region_change: {e}")

    def load_game_from_link(self, link):
        """Handler for clickable links (Game or Franchise)"""
        try:
            if ":" not in link or "|" not in link:
                # Fallback for old simple ID links
                game_id = int(link)
                self.selected_game_id = game_id
                self.status.setText(f"Переход к игре ID: {game_id}")
                self.fetch_details()
                return

            type_prefix, rest = link.split(":", 1)
            id_str, name_str = rest.split("|", 1)
            obj_id = int(id_str)

            if type_prefix == "game":
                self.selected_game_id = obj_id
                # Auto-fill name
                self._updating = True
                self.game_input.setText(name_str)
                self._updating = False
                self.status.setText(f"Переход к игре: {name_str}")
                self.fetch_details()
            
            elif type_prefix == "franchise":
                # Clear and Fill Franchise Input
                self._updating = True
                self.game_input.clear()
                self.franchise_input.setText(name_str)
                self._updating = False

                self.status.setText(f"Загрузка списка франшизы: {name_str}...")
                self.suggestions.clear()
                self.franchise_worker = FranchiseWorker(obj_id)
                self.franchise_worker.finished.connect(self.on_franchise_loaded)
                self.franchise_worker.error.connect(lambda e: self.status.setText(f"Ошибка франшизы: {e}"))
                self.franchise_worker.start()

        except Exception as e:
            print(f"Error parsing link: {e}")

    def search_franchises_manual(self):
        """Called by Franchise Search Button"""
        q = self.franchise_input.text().strip()
        if len(q) < 2:
            self.status.setText("Введите минимум 2 символа для франшизы")
            return
        
        self.status.setText(f"Поиск франшизы: {q}...")
        self.suggestions.clear()
        
        self.franchise_worker = FranchiseWorker(0, query=q, mode="search")
        self.franchise_worker.finished.connect(self.on_franchise_search_done)
        self.franchise_worker.error.connect(lambda e: self.status.setText(f"Ошибка: {e}"))
        self.franchise_worker.start()

    def on_franchise_search_done(self, results):
        self.suggestions.clear()
        self.navigation_stack = [] # Clear history on new search
        
        if not results:
            self.status.setText("Франшизы не найдены")
            return
        
        # If EXACTLY one franchise found, auto-load it
        if len(results) == 1:
            fid = results[0].get("id")
            name = results[0].get("name")
            self.status.setText(f"Открытие франшизы: {name}...")
            self.franchise_worker = FranchiseWorker(fid)
            self.franchise_worker.finished.connect(self.on_franchise_loaded)
            self.franchise_worker.error.connect(lambda e: self.status.setText(f"Ошибка франшизы: {e}"))
            self.franchise_worker.start()
            return

        # Else show list
        for f in results:
            name = f.get("name", "")
            fid = f.get("id")
            # Special Item for Franchise: ID and Type
            item = QListWidgetItem(f"📁 {name} (Франшиза)")
            item.setData(Qt.ItemDataRole.UserRole, fid)
            item.setData(Qt.ItemDataRole.UserRole + 1, "franchise") # Marker
            self.suggestions.addItem(item)
        self.status.setText(f"Найдено франшиз: {len(results)}")

    def on_franchise_loaded(self, games):
        """Called when games for a franchise are fetched (Step 2: Show Platforms)"""
        self.suggestions.clear()
        self.current_franchise_games = games 
        
        # Push current state to history if coming from somewhere else
        # Simple logic: if we just loaded franchise, we can go back to nothing or search results?
        # Let's handle "Back" button inside show_franchise_games_for_platform
        
        if not games:
            self.status.setText("Игры во франшизе не найдены")
            return
        
        platforms_map = {} 
        for g in games:
            plats = g.get("platforms", [])
            if plats:
                for pid in plats:
                    if pid not in platforms_map:
                        pname = igdb_client.get_platform_name_by_id(pid)
                        if pname: platforms_map[pid] = pname
                            
        # 1. Back Button (if history exists)
        if self.navigation_stack:
             back_item = QListWidgetItem("⬅ Назад")
             back_item.setData(Qt.ItemDataRole.UserRole, -999) # Back marker
             back_item.setData(Qt.ItemDataRole.UserRole + 1, "back_button")
             self.suggestions.addItem(back_item)

        # 2. All Platforms
        all_item = QListWidgetItem("📁 Все платформы")
        all_item.setData(Qt.ItemDataRole.UserRole, -1) 
        all_item.setData(Qt.ItemDataRole.UserRole + 1, "franchise_platform_folder")
        self.suggestions.addItem(all_item)
        
        # 3. Platform Folders
        sorted_platforms = sorted(platforms_map.items(), key=lambda x: x[1])
        for pid, pname in sorted_platforms:
            item = QListWidgetItem(f"📁 {pname}")
            item.setData(Qt.ItemDataRole.UserRole, pid)
            item.setData(Qt.ItemDataRole.UserRole + 1, "franchise_platform_folder")
            self.suggestions.addItem(item)
            
        self.status.setText(f"Выберите платформу ({len(sorted_platforms)} доступно)")

    def show_franchise_games_for_platform(self, platform_id):
        """Filter games by platform and display (Step 3: Show Games)"""
        # Save state before clearing
        self.navigation_stack.append("franchise_folders") # Marker for previous state
        
        self.suggestions.clear()
        
        # Back Button
        back_item = QListWidgetItem("⬅ Назад")
        back_item.setData(Qt.ItemDataRole.UserRole, -999)
        back_item.setData(Qt.ItemDataRole.UserRole + 1, "back_button")
        self.suggestions.addItem(back_item)
        
        filtered_games = []
        if platform_id == -1:
            filtered_games = self.current_franchise_games
        else:
            for g in self.current_franchise_games:
                if platform_id in g.get("platforms", []):
                    filtered_games.append(g)
        
        for g in filtered_games:
            name = g.get("name", "")
            gid = g.get("id")
            platforms = g.get("platforms", []) 
            
            if platform_id != -1:
                pname = igdb_client.get_platform_name_by_id(platform_id)
                display_text = f"{name} ({pname})"
                item = QListWidgetItem(display_text)
                item.setData(Qt.ItemDataRole.UserRole, gid)
                item.setData(Qt.ItemDataRole.UserRole + 1, [platform_id])
                item.setData(Qt.ItemDataRole.UserRole + 2, name)
                self.suggestions.addItem(item)
            else:
                if not platforms:
                    item = QListWidgetItem(name)
                    item.setData(Qt.ItemDataRole.UserRole, gid)
                    item.setData(Qt.ItemDataRole.UserRole + 1, [])
                    item.setData(Qt.ItemDataRole.UserRole + 2, name)
                    self.suggestions.addItem(item)
                else:
                    for pid in platforms:
                        pname = igdb_client.get_platform_name_by_id(pid)
                        if pname:
                            display_text = f"{name} ({pname})"
                            item = QListWidgetItem(display_text)
                            item.setData(Qt.ItemDataRole.UserRole, gid)
                            item.setData(Qt.ItemDataRole.UserRole + 1, [pid])
                            item.setData(Qt.ItemDataRole.UserRole + 2, name)
                            self.suggestions.addItem(item)

        self.status.setText(f"Игр: {self.suggestions.count() - 1}") # -1 for back button

    def clear_all(self):
        """Clear all inputs and results"""
        self._updating = True
        self.game_input.clear()
        self.franchise_input.clear() # Added franchise input clear
        self.suggestions.clear()
        self.search_results = []
        self.selected_game_id = None
        self.search_btn.setEnabled(False)
        self.current_franchise_games = []
        self.navigation_stack = [] # Clear history
        
        # Clear results area
        for lbl in [self.name_res, self.date_res, self.status_res, self.genres_res, 
                    self.regions_res, self.ru_res, self.players_res, self.devs_res, 
                    self.pubs_res, self.franchise_res, self.parent_res, self.dlc_res, 
                    self.remake_res]:
            lbl.setText("—")
            lbl.setToolTip("")
        
        self.ra_res.setText("—")
        self.ra_res.setToolTip("")
        self.ra_list_btn.setEnabled(False)
        self.desc_text.clear()
        
        self.cover_label.clear()
        self.cover_label.setText("Нет изображения")
        self.cover_pixmap = None
        self.cover_label.set_image_path(None)
        
        for lbl in self.ss_labels:
            lbl.clear()
            lbl.set_image_path(None)
            
        self.ss_counter.setText("0 / 0")
        self.prev_ss.setEnabled(False)
        self.next_ss.setEnabled(False)
        
        self._updating = False
        self.status.setText("Очищено")

    def on_game_input(self, t):
        try:
            if self._updating: return
            self.search_timer.stop()
            self.suggestions.clear(); self.search_results = []; self.selected_game_id = None; self.search_btn.setEnabled(False)
            if len(t) >= 2: self.search_timer.start(300); self.status.setText("Ввод...")
            else: self.status.setText("Введите минимум 2 символа")
        except Exception as e: print(f"Error in on_game_input: {e}")
    
    def perform_search(self):
        try:
            q = self.game_input.text().strip()
            if len(q) < 2: return
            pid = igdb_client.get_platform_id(self.platform_combo.currentText())
            
            source = "igdb"
            if self.rb_moby.isChecked():
                source = "moby"
            
            if source == "igdb" and not pid: 
                self.status.setText("Платформа не найдена"); return
            
            self.status.setText(f"🔄 Поиск ({source})...")
            self.search_worker = SearchWorker(q, pid, None, source=source)
            self.search_worker.finished.connect(self.on_search_done)
            self.search_worker.error.connect(lambda e: self.status.setText(f"Ошибка: {e}"))
            self.search_worker.start()
        except Exception as e: print(f"Error in perform_search: {e}"); self.status.setText(f"Ошибка: {e}")
    
    def on_search_done(self, results):
        self.search_results = results; self.suggestions.clear()
        if not results: self.status.setText("Не найдено"); return
        for g in results:
            item = QListWidgetItem(g.get("name", ""))
            item.setData(Qt.ItemDataRole.UserRole, g.get("id"))
            # For search results, we store platforms if available (added to search query in igdb_api)
            item.setData(Qt.ItemDataRole.UserRole + 1, g.get("platforms", []))
            item.setData(Qt.ItemDataRole.UserRole + 2, g.get("name", "")) # Real Name
            self.suggestions.addItem(item)
        self.status.setText(f"Найдено: {len(results)}")
    
    def on_suggestion_click(self, item):
        try:
            # Check if it is a franchise item
            marker = item.data(Qt.ItemDataRole.UserRole + 1)
            
            # Case 1: Root Franchise Item (from Search)
            if marker == "franchise":
                fid = item.data(Qt.ItemDataRole.UserRole)
                self.status.setText(f"Загрузка игр франшизы ID: {fid}...")
                self.suggestions.clear()
                self.franchise_worker = FranchiseWorker(fid)
                self.franchise_worker.finished.connect(self.on_franchise_loaded)
                self.franchise_worker.error.connect(lambda e: self.status.setText(f"Ошибка франшизы: {e}"))
                self.franchise_worker.start()
                return

            # Case 2: Platform Folder inside Franchise
            if marker == "franchise_platform_folder":
                pid = item.data(Qt.ItemDataRole.UserRole)
                self.show_franchise_games_for_platform(pid)
                return

            # Case 2.1: Back Button
            if marker == "back_button":
                if self.navigation_stack:
                    state = self.navigation_stack.pop()
                    if state == "franchise_folders":
                        # Reload folder view
                        self.on_franchise_loaded(self.current_franchise_games)
                return

            # Case 3: Actual Game
            self.selected_game_id = item.data(Qt.ItemDataRole.UserRole)
            game_platforms = item.data(Qt.ItemDataRole.UserRole + 1) # List of IDs
            # Use stored real name if available (clean of platform text), else item text
            game_name = item.data(Qt.ItemDataRole.UserRole + 2) or item.text()
            
            self._updating = True
            
            # 1. Update Name
            self.game_input.setText(game_name)
            
            # 2. Update Platform if needed
            current_pid = igdb_client.get_platform_id(self.platform_combo.currentText())
            
            # Logic: If current platform is NOT in game's platform list, switch to the first available one.
            if game_platforms and isinstance(game_platforms, list):
                if current_pid not in game_platforms:
                    # Try to find the name of the first platform ID
                    first_pid = game_platforms[0]
                    p_name = igdb_client.get_platform_name_by_id(first_pid)
                    if p_name:
                        index = self.platform_combo.findText(p_name)
                        if index >= 0:
                            self.platform_combo.setCurrentIndex(index)
                            self.platform_id = first_pid # Sync
                            self.status.setText(f"Платформа переключена на: {p_name}")

            self._updating = False
            self.search_btn.setEnabled(True)
            self.status.setText(f"Выбрано: {game_name}")
        except Exception as e: 
            print(f"Error clicking suggestion: {e}")
            self._updating = False
    
    def on_suggestion_dblclick(self, item): 
        # Only fetch details if it's a game (not franchise or folder)
        marker = item.data(Qt.ItemDataRole.UserRole + 1)
        if marker != "franchise" and marker != "franchise_platform_folder" and marker != "back_button":
            self.on_suggestion_click(item)
            self.fetch_details()
    
    def show_all_games(self):
        pn = self.platform_combo.currentText(); pid = igdb_client.get_platform_id(pn)
        if not pid: QMessageBox.warning(self, "Внимание", "Выберите платформу"); return
        
        # Определяем текущий источник
        source = "moby" if self.rb_moby.isChecked() else "igdb"
        
        if not self.all_games_dialog: self.all_games_dialog = AllGamesDialog(self); self.all_games_dialog.game_selected.connect(self.on_game_from_dialog)
        self.all_games_dialog.load_games(pn, pid, source); self.all_games_dialog.show()
    
    def on_game_from_dialog(self, name, gid):
        self._updating = True; self.game_input.setText(name); self._updating = False; self.selected_game_id = gid; 
        self.suggestions.clear(); self.search_btn.setEnabled(True); self.fetch_details()
    
    def force_refresh(self):
        if not self.selected_game_id:
            QMessageBox.information(self, "Инфо", "Сначала выберите игру для обновления")
            return
            
        # Prevent crash from double clicking
        self.refresh_btn.setEnabled(False)
        self.status.setText("Удаление кэша...")
        QApplication.processEvents()

        source = "moby" if self.rb_moby.isChecked() else "igdb"
        
        pid = self.platform_id if source == "igdb" else "None" 
        if source == "igdb" and not pid:
             pid = igdb_client.get_platform_id(self.platform_combo.currentText())
        
        path = os.path.join(CacheManager.GAMES_DIR, f"{self.selected_game_id}_{pid}_{source}.json")
        
        # Try clean removal
        try:
            if os.path.exists(path):
                os.remove(path)
        except Exception as e:
            print(f"Error removing cache: {e}")
        
        self.status.setText("Обновление...")
        self.fetch_details()
        
        # Re-enable after a moment (fetch_details is async)
        QTimer.singleShot(2000, lambda: self.refresh_btn.setEnabled(True))

    def fetch_details(self):
        if not self.selected_game_id: QMessageBox.warning(self, "Внимание", "Выберите игру"); return
        
        source = "igdb"
        if self.rb_moby.isChecked():
            source = "moby"
            
        self.status.setText("🔄 Загрузка..."); self.search_btn.setEnabled(False)
        
        if self.details_worker and self.details_worker.isRunning():
            self.details_worker.wait() # Ensure previous is done
            
        self.details_worker = DetailsWorker(self.selected_game_id, self.platform_id, None, source=source)
        
        self.details_worker.finished.connect(self.on_details_done)
        self.details_worker.error.connect(self.on_details_err)
        self.details_worker.start()
    
    def on_details_done(self, d):
        try:
            self.current_game_details = d; self.name_res.setText(d.get("name", "—")); 
            
            rd = d.get("earliest_release_date")
            if isinstance(rd, (int, float)):
                self.date_res.setText(igdb_client.format_release_date(rd))
            else:
                self.date_res.setText(str(rd) if rd else "Неизвестно")

            # Status
            self.status_res.setText(d.get("status_display", "—"))
            
            # Russian
            self.ru_res.setText(d.get("russian_support", "None"))

            mixed_g = d.get("mixed_genres", "")
            if not mixed_g:
                genres_data = d.get("genres")
                if isinstance(genres_data, list):
                    mixed_g = igdb_client.format_genres(genres_data)
                else:
                    mixed_g = str(genres_data)
            self.genres_res.setText(mixed_g)

            regions_list = d.get("all_regions", [])
            self.regions_res.setText("; ".join(regions_list) if regions_list else "—")

            self.players_res.setText(str(d.get("max_local_players", 1)))
            
            devs = d.get("developers_list")
            self.devs_res.setText(igdb_client.format_list(devs) if isinstance(devs, list) else str(devs))
            
            pubs = d.get("publishers_list")
            self.pubs_res.setText(igdb_client.format_list(pubs) if isinstance(pubs, list) else str(pubs))
            
            # Relations (Formatting Links)
            def format_links(data_list):
                if not data_list: return "—"
                links = []
                for item in data_list:
                    # Create HTML link with ID
                    links.append(f"<a href='{item['id']}' style='color: #00d4ff; text-decoration: none;'>{item['name']}</a>")
                return "; ".join(links)

            self.franchise_res.setText(d.get("franchise_link_str", "—"))
            self.parent_res.setText(d.get("parent_link_str", "—"))
            
            # Tooltip handling for long lists
            dlc_links = d.get("dlc_link_str", "—")
            self.dlc_res.setText(dlc_links)
            if len(d.get("dlcs_str", "")) > 100:
                self.dlc_res.setToolTip(d.get("dlcs_str"))
            
            remake_links = d.get("remake_link_str", "—")
            self.remake_res.setText(remake_links)
            if len(d.get("remakes_remasters_str", "")) > 100:
                self.remake_res.setToolTip(d.get("remakes_remasters_str"))

            self.desc_text.setText(d.get("description") or "Нет описания")
            
            cover = d.get("cover_url_large")
            if cover: self.load_img(cover, "cover", API_CONFIG.get("CACHE_SEARCH_COVER", True))
            else: 
                self.cover_label.setText("Нет обложки")
                self.cover_label.setPixmap(QPixmap())
                self.cover_label.set_image_path(None)
                self.cover_pixmap = None
            
            self.current_ss_page = 0; self.update_ss(); self.search_btn.setEnabled(True); self.status.setText("✅ Загружено")

            self.ra_res.setText("Поиск...")
            self.ra_thread = RAThread(d.get("name"), self.platform_id)
            self.ra_thread.finished.connect(self.on_ra_finished)
            self.ra_thread.start()
            
        except Exception as e: print(e); self.search_btn.setEnabled(True)

    def on_ra_finished(self, info):
        self.current_ra_info = info
        if info:
            try: ach = int(info.get("NumAchievements", 0))
            except: ach = 0
            
            if ach == 0 and "Achievements" in info:
                ach = len(info.get("Achievements", {}))
                
            gid = info.get("ID")
            title = info.get("Title", "Unknown")
            console_name = info.get("ConsoleName", "Unknown")
            
            url = f"https://retroachievements.org/game/{gid}"
            self.ra_res.setText(f"<a href='{url}' style='color: #ffcc00; text-decoration: none; font-weight: bold;'>{ach} 🏆</a>")
            self.ra_res.setToolTip(f"Найдено RA: {title} (ID: {gid})\nПлатформа: {console_name}")
            self.ra_list_btn.setEnabled(True)
        else:
            self.ra_res.setText("—")
            self.ra_res.setToolTip("Игра не найдена в базе RetroAchievements")
            self.ra_list_btn.setEnabled(False)

    def show_ra_list(self):
        if not self.current_ra_info: return
        achs = self.current_ra_info.get("Achievements", {})
        title = self.current_ra_info.get("Title", "Unknown")
        dlg = AchievementsDialog(self, achs, title)
        dlg.exec()
    
    def on_details_err(self, e): self.status.setText(f"Ошибка: {e}"); self.search_btn.setEnabled(True); QMessageBox.critical(self, "Ошибка", e)
    
    def load_img(self, url, t, save_cache=True):
        loader = ImageLoader(url, t, save_to_cache=save_cache)
        loader.finished.connect(self.on_img_loaded)
        loader.start()
        self.image_loaders.append(loader)
    
    def on_img_loaded(self, pix, t):
        # Calculate correct path for context menu
        # We need to know which URL was loaded. The loader has self.url
        sender = self.sender()
        if not sender: return
        
        path = CacheManager.get_image_path(sender.url) if sender.url else None

        if t == "cover" and not pix.isNull():
            self.cover_pixmap = pix
            scaled = pix.scaled(self.cover_label.width()-10, self.cover_label.height()-10, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
            self.cover_label.setPixmap(scaled)
            self.cover_label.set_image_path(path)
            
        elif t.startswith("screenshot_"):
            try:
                i = int(t.split("_")[1])
                if i < len(self.ss_labels):
                    self.ss_labels[i].setPixmap(pix.scaled(self.ss_labels[i].size(), Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                    self.ss_labels[i].set_image_path(path)
            except: pass
    
    def show_cover_fullscreen(self, _):
        if not self.cover_pixmap or self.cover_pixmap.isNull(): return
        if not self.cover_viewer: self.cover_viewer = ImageViewer(self)
        self.cover_viewer.show_pixmap(self.cover_pixmap); self.cover_viewer.setGeometry(self.geometry()); self.cover_viewer.show()
    
    def update_ss(self):
        if not self.current_game_details: return
        
        # Check cache setting for SS
        save_ss = API_CONFIG.get("CACHE_SEARCH_SS", True)
        
        ss = self.current_game_details.get("screenshot_urls", []); total = len(ss)
        for lbl in self.ss_labels: 
            lbl.clear()
            lbl.set_image_path(None)
            
        if total == 0: self.ss_counter.setText("0 / 0"); self.prev_ss.setEnabled(False); self.next_ss.setEnabled(False); return
        pages = (total + self.ss_per_page - 1) // self.ss_per_page; start = self.current_ss_page * self.ss_per_page
        for i in range(self.ss_per_page):
            idx = start + i
            if idx < total: url = ss[idx].get("thumb", ""); self.load_img(url, f"screenshot_{i}", save_ss) if url else None
        self.ss_counter.setText(f"{self.current_ss_page + 1} / {pages}"); self.prev_ss.setEnabled(self.current_ss_page > 0); self.next_ss.setEnabled(self.current_ss_page < pages - 1)
    
    def prev_ss_page(self):
        if self.current_ss_page > 0: self.current_ss_page -= 1; self.update_ss()
    
    def next_ss_page(self):
        if self.current_game_details:
            ss = self.current_game_details.get("screenshot_urls", []); pages = (len(ss) + self.ss_per_page - 1) // self.ss_per_page
            if self.current_ss_page < pages - 1: self.current_ss_page += 1; self.update_ss()
    
    def on_ss_clicked(self, i):
        if not self.current_game_details: return
        ss = self.current_game_details.get("screenshot_urls", [])
        if not ss: return
        idx = self.current_ss_page * self.ss_per_page + i
        if idx >= len(ss): return
        if not self.ss_viewer: self.ss_viewer = ScreenshotViewer(self)
        self.ss_viewer.set_screenshots(ss, idx); self.ss_viewer.setGeometry(self.geometry()); self.ss_viewer.show()


def main():
    app = QApplication(sys.argv); app.setStyleSheet(DARK_STYLESHEET)
    win = MainWindow(); win.show(); sys.exit(app.exec())

if __name__ == "__main__": main()